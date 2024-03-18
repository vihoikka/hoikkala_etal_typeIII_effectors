import pandas as pd
import argparse
import os

# Create args for mastertable and outputfile
parser = argparse.ArgumentParser(description='Adds hyperlinks to the mastertable')
parser.add_argument('-m', '--mastertable', help='mastertable', required=True)
parser.add_argument('-vf', '--viz_dir_full', help='viz_dir full', required=True)
parser.add_argument('-vr', '--viz_dir_relative', help='viz_dir rel', required=True)
parser.add_argument('-o', '--output_excel', help='output_excel', required=True)

args = parser.parse_args()
mastertable_path = args.mastertable
viz_dir_full = args.viz_dir_full
viz_dir_relative = args.viz_dir_relative
output_excel = args.output_excel

# Read the .tsv file into a DataFrame
mastertable = pd.read_csv(mastertable_path, sep='\t')

# Add a new column with the file paths to the images
mastertable['Visualisation_path'] = viz_dir_relative + "/" + mastertable['Locus'] + '/' + mastertable['Locus'] + '_viz.png'
mastertable['Visualisation_table_path'] = viz_dir_relative + "/" + mastertable['Locus'] + '/' + mastertable['Locus'] + '_merged.csv'

#create new column Visualisation_link but leave it blank for now
mastertable['Visualisation_link'] = ""

# Create a HTML file for each image that displays the image
for index, row in mastertable.iterrows():
    html_path = os.path.join(viz_dir_full, row['Locus'], row['Locus'] + '.html')
    html_path_relative = os.path.join(viz_dir_relative, row['Locus'], row['Locus'] + '.html')
    mastertable.loc[index, "Visualisation_link"] = html_path_relative
    with open(html_path, 'w') as file:
        file.write(f'<img src="{row["Locus"]}_viz.png">')


# Convert the file paths to excel hyperlinks
def path_to_hyperlink(path):
    return f'=HYPERLINK("{path}", " "Open viz")'

#mastertable['Visualisation_link'] = mastertable['Visualisation_link'].apply(path_to_hyperlink)

#rearrange columns in the dataframe by making Locus first, and Visualisation_link second
# cols = mastertable.columns.tolist()
# cols = cols[-1:] + cols[:-1]
# cols = cols[-1:] + cols[:-1]
# mastertable = mastertable[cols]

# Write the DataFrame to an Excel file
mastertable.to_excel(output_excel, index = False, engine='openpyxl')

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Load your workbook and select a sheet
workbook = load_workbook(filename=output_excel)
sheet = workbook.active

# Find the hyperlink column (assuming it's the last one)
hyperlink_col = sheet.max_column

# Iterate over your data and add Excel hyperlinks
for row in range(2, sheet.max_row+1):   # Will start from row 2 since headers are at row 1
    cell = sheet.cell(row=row, column=hyperlink_col)
    link = cell.value    # Get value added by Pandas from the cell
    cell.value = '=HYPERLINK("%s", "Open viz")' % link

# Save the workbook
workbook.save(filename=output_excel)
workbook.close()